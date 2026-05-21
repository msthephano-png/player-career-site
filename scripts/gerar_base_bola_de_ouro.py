import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE_JS = ROOT / "assets" / "js" / "bola-de-ouro.js"
OUTPUT_XLSX = ROOT / "data" / "bola_de_ouro_ate_2050.xlsx"
OUTPUT_JSON = ROOT / "data" / "bola_de_ouro_ate_2050.json"


def extract_ballon_data():
    content = SOURCE_JS.read_text(encoding="utf-8")
    match = re.search(r"const ballonData = \{(?P<body>.*?)\n\};", content, re.S)
    if not match:
        raise RuntimeError("Nao foi possivel localizar o objeto ballonData.")

    body = match.group("body")
    data = {}
    for year, raw_list in re.findall(r"(\d{4}):\s*(\[[^\]]*\]|null)", body, re.S):
        year_int = int(year)
        if raw_list == "null" or year_int > 2050:
            continue
        names = re.findall(r'"([^"]*)"', raw_list)
        data[year_int] = names

    return dict(sorted(data.items()))


def build_rows(ballon_data):
    ranking_rows = []
    stats = defaultdict(lambda: {"indicacoes": 0, "vitorias": 0, "vices": 0, "terceiros": 0, "podios": 0, "posicoes": []})

    for year, names in ballon_data.items():
        for index, player in enumerate(names, start=1):
            podium = index <= 3
            medal = "Ouro" if index == 1 else "Prata" if index == 2 else "Bronze" if index == 3 else ""

            ranking_rows.append(
                {
                    "Ano": year,
                    "Posição": index,
                    "Jogador": player,
                    "Pódio": "Sim" if podium else "Não",
                    "Medalha": medal,
                }
            )

            stats[player]["indicacoes"] += 1
            stats[player]["posicoes"].append({"ano": year, "posicao": index})
            if index == 1:
                stats[player]["vitorias"] += 1
                stats[player]["podios"] += 1
            elif index == 2:
                stats[player]["vices"] += 1
                stats[player]["podios"] += 1
            elif index == 3:
                stats[player]["terceiros"] += 1
                stats[player]["podios"] += 1

    player_rows = []
    for player, values in stats.items():
        positions = [item["posicao"] for item in values["posicoes"]]
        years = [item["ano"] for item in values["posicoes"]]
        best = min(positions)
        best_years = [str(item["ano"]) for item in values["posicoes"] if item["posicao"] == best]
        history = "; ".join(f'{item["ano"]}: {item["posicao"]}º' for item in values["posicoes"])

        player_rows.append(
            {
                "Jogador": player,
                "Indicações": values["indicacoes"],
                "Pódios": values["podios"],
                "Vitórias": values["vitorias"],
                "Vices": values["vices"],
                "Terceiros": values["terceiros"],
                "Melhor posição": best,
                "Ano(s) da melhor posição": ", ".join(best_years),
                "Primeira indicação": min(years),
                "Última indicação": max(years),
                "Histórico": history,
            }
        )

    player_rows.sort(
        key=lambda row: (
            -row["Vitórias"],
            -row["Pódios"],
            -row["Indicações"],
            row["Melhor posição"],
            row["Jogador"],
        )
    )

    return ranking_rows, player_rows


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        width = min(max(max_len + 3, 12), 72)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_workbook(ranking_rows, player_rows, ballon_data):
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Resumo"
    ws_summary.append(["Indicador", "Valor"])
    ws_summary.append(["Edições fechadas", len(ballon_data)])
    ws_summary.append(["Período", f"{min(ballon_data)}-{max(ballon_data)}"])
    ws_summary.append(["Linhas de ranking", len(ranking_rows)])
    ws_summary.append(["Jogadores únicos", len(player_rows)])
    ws_summary.append(["Fonte principal", "assets/js/bola-de-ouro.js"])

    ws_rankings = wb.create_sheet("Rankings Anuais")
    ws_rankings.append(["Ano", "Posição", "Jogador", "Pódio", "Medalha"])
    for row in ranking_rows:
        ws_rankings.append([row["Ano"], row["Posição"], row["Jogador"], row["Pódio"], row["Medalha"]])

    ws_players = wb.create_sheet("Estatísticas por Jogador")
    headers = [
        "Jogador",
        "Indicações",
        "Pódios",
        "Vitórias",
        "Vices",
        "Terceiros",
        "Melhor posição",
        "Ano(s) da melhor posição",
        "Primeira indicação",
        "Última indicação",
        "Histórico",
    ]
    ws_players.append(headers)
    for row in player_rows:
        ws_players.append([row[header] for header in headers])

    for ws in wb.worksheets:
        style_sheet(ws)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)


def write_json(ranking_rows, player_rows):
    payload = {
        "rankings_anuais": ranking_rows,
        "estatisticas_por_jogador": player_rows,
    }
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main():
    ballon_data = extract_ballon_data()
    ranking_rows, player_rows = build_rows(ballon_data)
    write_workbook(ranking_rows, player_rows, ballon_data)
    write_json(ranking_rows, player_rows)
    print(OUTPUT_XLSX)
    print(OUTPUT_JSON)


if __name__ == "__main__":
    main()
