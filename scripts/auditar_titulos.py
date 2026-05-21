import json
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "data" / "campeoes_completo_ate_2050.xlsx"
JSON_PATH = ROOT / "assets" / "data" / "competicoes.json"


def sheet_rows(workbook, sheet_name):
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue
        rows.append(dict(zip(headers, row)))
    return rows


def normalize(value):
    return str(value or "").strip().lower()


def main():
    workbook = load_workbook(XLSX, data_only=True)
    club_rows = sheet_rows(workbook, "Títulos por Clube")
    rank_rows = sheet_rows(workbook, "Ranking por Competição")
    base_rows = sheet_rows(workbook, "Base Consolidada")

    payload = json.loads(JSON_PATH.read_text(encoding="utf-8"))

    print("=== ABAS DA PLANILHA ===")
    print(", ".join(workbook.sheetnames))

    focus_teams = ["real madrid", "real", "barcelona", "bayern", "psg", "milan"]
    focus_comps = ["champions league", "la liga", "laliga", "brasileirao", "brasileirão"]

    print("\n=== PLANILHA | Títulos por Clube (Champions / Real) ===")
    for row in club_rows:
        clube = normalize(row.get("Clube/Seleção"))
        titulo = normalize(row.get("Título"))
        if "real" in clube and "champion" in titulo:
            print(
                f"  {row.get('Clube/Seleção')} | {row.get('Título')} | "
                f"total={row.get('Total')} | fonte={row.get('Fonte')}"
            )

    print("\n=== PLANILHA | Ranking por Competição (Champions) ===")
    for row in rank_rows:
        comp = normalize(row.get("Competição"))
        if "champion" in comp and "conference" not in comp and "europa" not in comp:
            print(f"\n  Competição: {row.get('Competição')}")
            champions = [
                item
                for item in rank_rows
                if item.get("Competição") == row.get("Competição")
            ]
            for item in sorted(champions, key=lambda x: -(x.get("Títulos") or 0))[:8]:
                print(
                    f"    {item.get('Campeão')}: {item.get('Títulos')} "
                    f"({item.get('Primeiro ano')}-{item.get('Último ano')})"
                )
            break

    print("\n=== PLANILHA | Real no ranking Champions ===")
    for row in rank_rows:
        if normalize(row.get("Competição")) == "champions league" and "real" in normalize(
            row.get("Campeão")
        ):
            print(
                f"  {row.get('Campeão')}: {row.get('Títulos')} títulos "
                f"({row.get('Primeiro ano')}-{row.get('Último ano')})"
            )

    print("\n=== PLANILHA | Vitórias na simulação (Base Consolidada, Champions) ===")
    sim_by_team = Counter()
    for row in base_rows:
        comp = normalize(row.get("Competição"))
        if row.get("Origem") == "Site / simulação" and comp == "champions league":
            sim_by_team[row.get("Campeão")] += 1
    for team, total in sim_by_team.most_common(12):
        print(f"  {team}: {total}")

    print("\n=== JSON | competicoes.json ===")
    champions_comp = next(
        (comp for comp in payload["competitions"] if comp["name"] == "Champions League"),
        None,
    )
    if champions_comp:
        sim_results = [
            item
            for item in champions_comp["results"]
            if item.get("source") == "Site / simulação"
        ]
        json_sim = Counter(item["champion"] for item in sim_results)
        print(f"  Resultados simulação: {len(sim_results)} edições")
        print("  Vitórias por time (sim):")
        for team, total in json_sim.most_common(12):
            print(f"    {team}: {total}")
        print("  Ranking JSON (top 8):")
        for item in champions_comp["ranking"][:8]:
            print(f"    {item['team']}: {item['titles']}")

    print("\n=== JSON | clubRelevantTitles (Real + Champions) ===")
    for entry in payload.get("clubRelevantTitles", []):
        if "real" in normalize(entry["team"]):
            for title in entry["titles"]:
                if "champion" in normalize(title["title"]):
                    print(
                        f"  {entry['team']} | {title['title']} | "
                        f"{title['total']}x | {title.get('source')}"
                    )

    print("\n=== AUDITORIA | Consistência planilha x JSON x site ===")
    issues = []

    # Compare ranking sheet vs JSON for Champions
    sheet_champions = [
        row
        for row in rank_rows
        if normalize(row.get("Competição")) == "champions league"
    ]
    json_rank = {
        item["team"]: item["titles"]
        for item in (champions_comp or {}).get("ranking", [])
    }
    sheet_rank = {row.get("Campeão"): row.get("Títulos") for row in sheet_champions}

    all_teams = set(sheet_rank) | set(json_rank)
    for team in sorted(all_teams):
        sheet_total = sheet_rank.get(team)
        json_total = json_rank.get(team)
        if sheet_total != json_total:
            issues.append(
                f"Ranking Champions diverge: {team} planilha={sheet_total} json={json_total}"
            )

    # Compare club titles sheet vs JSON clubRelevantTitles
    json_club = {}
    for entry in payload.get("clubRelevantTitles", []):
        json_club[entry["team"]] = {
            item["title"]: item["total"] for item in entry["titles"]
        }

    for row in club_rows:
        team = row.get("Clube/Seleção")
        title = row.get("Título")
        sheet_total = row.get("Total")
        json_total = json_club.get(team, {}).get(title)
        if json_total is None:
            if "champion" in normalize(title) and "real" in normalize(team):
                issues.append(
                    f"Clube na planilha sem entrada JSON: {team} | {title} = {sheet_total}"
                )
        elif sheet_total != json_total:
            issues.append(
                f"Títulos por clube divergem: {team} | {title} planilha={sheet_total} json={json_total}"
            )

    # Real Madrid Champions specifically
    real_sheet_rank = next(
        (
            row.get("Títulos")
            for row in sheet_champions
            if normalize(row.get("Campeão")) == "real madrid"
        ),
        None,
    )
    real_json_rank = json_rank.get("Real Madrid")
    real_club_sheet = next(
        (
            row.get("Total")
            for row in club_rows
            if normalize(row.get("Clube/Seleção")) == "real madrid"
            and normalize(row.get("Título")) == "champions league"
        ),
        None,
    )
    real_club_json = json_club.get("Real Madrid", {}).get("Champions League")

    print(f"  Real Madrid Champions:")
    print(f"    Ranking planilha: {real_sheet_rank}")
    print(f"    Ranking JSON:     {real_json_rank}")
    print(f"    Títulos/clube planilha: {real_club_sheet}")
    print(f"    Títulos/clube JSON:     {real_club_json}")

    # dados.js embedded teamData for Real (from grep earlier - LaLiga etc)
    print("\n=== SITE (dados.js fallback) | Real Madrid titles in teamData ===")
    print("  (Champions total no modal vem do JSON clubRelevantTitles após fetch)")

    if issues:
        print(f"\n  Encontradas {len(issues)} divergências:")
        for issue in issues[:40]:
            print(f"    - {issue}")
        if len(issues) > 40:
            print(f"    ... e mais {len(issues) - 40}")
    else:
        print("\n  Nenhuma divergência entre planilha Títulos/Ranking e JSON exportado.")

    # Broader mismatch scan top 20 club/title pairs
    print("\n=== TOP divergências Títulos por Clube (planilha vs JSON) ===")
    mismatches = []
    for row in club_rows:
        team = row.get("Clube/Seleção")
        title = row.get("Título")
        sheet_total = row.get("Total")
        json_total = json_club.get(team, {}).get(title)
        if json_total is not None and sheet_total != json_total:
            mismatches.append((team, title, sheet_total, json_total))
    mismatches.sort(key=lambda x: abs((x[2] or 0) - (x[3] or 0)), reverse=True)
    for team, title, sheet_total, json_total in mismatches[:25]:
        print(f"  {team} | {title}: planilha={sheet_total} json={json_total}")


if __name__ == "__main__":
    main()
