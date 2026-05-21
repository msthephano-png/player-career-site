import json
import sqlite3
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
DB_PATH = DATA_DIR / "modo_carreira.sqlite"
CHAMPIONS_XLSX = DATA_DIR / "campeoes_completo_ate_2050.xlsx"
PLAYERS_JSON = DATA_DIR / "jogadores_carreiras.json"
BALLON_JSON = DATA_DIR / "bola_de_ouro_ate_2050.json"


SCHEMA = """
PRAGMA foreign_keys = ON;

DROP TABLE IF EXISTS championship_results;
DROP TABLE IF EXISTS competition_champion_rankings;
DROP TABLE IF EXISTS club_relevant_titles;
DROP TABLE IF EXISTS ballon_rankings;
DROP TABLE IF EXISTS ballon_player_stats;
DROP TABLE IF EXISTS player_competition_stats;
DROP TABLE IF EXISTS player_spell_honors;
DROP TABLE IF EXISTS player_aggregate_honors;
DROP TABLE IF EXISTS player_spells;
DROP TABLE IF EXISTS players;

CREATE TABLE players (
  slug TEXT PRIMARY KEY,
  name TEXT NOT NULL,
  matches INTEGER NOT NULL,
  goals INTEGER NOT NULL,
  assists INTEGER NOT NULL,
  hub_matches INTEGER,
  hub_goals INTEGER,
  hub_assists INTEGER,
  hub_matches_diff INTEGER,
  hub_goals_diff INTEGER,
  hub_assists_diff INTEGER,
  collective_titles INTEGER,
  individual_awards INTEGER
);

CREATE TABLE player_spells (
  id TEXT PRIMARY KEY,
  player_slug TEXT NOT NULL REFERENCES players(slug),
  player_name TEXT NOT NULL,
  spell_order INTEGER NOT NULL,
  team_name TEXT NOT NULL,
  team_full_name TEXT NOT NULL,
  team_type TEXT NOT NULL,
  matches INTEGER NOT NULL,
  goals INTEGER NOT NULL,
  assists INTEGER NOT NULL,
  page_path TEXT,
  description TEXT
);

CREATE TABLE player_competition_stats (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  spell_id TEXT NOT NULL REFERENCES player_spells(id),
  player_name TEXT NOT NULL,
  team_name TEXT NOT NULL,
  competition_name TEXT NOT NULL,
  matches INTEGER NOT NULL,
  goals INTEGER NOT NULL,
  assists INTEGER NOT NULL
);

CREATE TABLE player_spell_honors (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  spell_id TEXT NOT NULL REFERENCES player_spells(id),
  player_name TEXT NOT NULL,
  team_name TEXT NOT NULL,
  category TEXT NOT NULL,
  honor_name TEXT NOT NULL,
  quantity INTEGER NOT NULL
);

CREATE TABLE player_aggregate_honors (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  player_slug TEXT NOT NULL REFERENCES players(slug),
  player_name TEXT NOT NULL,
  category TEXT NOT NULL,
  honor_name TEXT NOT NULL,
  quantity INTEGER NOT NULL
);

CREATE TABLE championship_results (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  category TEXT NOT NULL,
  type TEXT NOT NULL,
  competition_name TEXT NOT NULL,
  season TEXT NOT NULL,
  champion_name TEXT NOT NULL,
  source TEXT NOT NULL
);

CREATE TABLE competition_champion_rankings (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  competition_name TEXT NOT NULL,
  champion_name TEXT NOT NULL,
  titles INTEGER NOT NULL,
  first_year INTEGER,
  last_year INTEGER
);

CREATE TABLE club_relevant_titles (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  team_name TEXT NOT NULL,
  title_name TEXT NOT NULL,
  total INTEGER NOT NULL,
  source TEXT NOT NULL
);

CREATE TABLE ballon_rankings (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  year INTEGER NOT NULL,
  position INTEGER NOT NULL,
  player_name TEXT NOT NULL,
  podium TEXT NOT NULL,
  medal TEXT,
  UNIQUE(year, position)
);

CREATE TABLE ballon_player_stats (
  player_name TEXT PRIMARY KEY,
  nominations INTEGER NOT NULL,
  podiums INTEGER NOT NULL,
  wins INTEGER NOT NULL,
  seconds INTEGER NOT NULL,
  thirds INTEGER NOT NULL,
  best_position INTEGER NOT NULL,
  best_position_years TEXT,
  first_nomination INTEGER,
  last_nomination INTEGER,
  history TEXT
);

CREATE INDEX idx_championship_results_competition ON championship_results(competition_name, season);
CREATE INDEX idx_championship_results_champion ON championship_results(champion_name);
CREATE INDEX idx_player_spells_player ON player_spells(player_slug);
CREATE INDEX idx_player_competition_stats_spell ON player_competition_stats(spell_id);
CREATE INDEX idx_ballon_rankings_year ON ballon_rankings(year);
"""


def load_json(path):
    with path.open("r", encoding="utf-8") as file:
        return json.load(file)


def rows_from_sheet(workbook, sheet_name):
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[1]]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(value is not None for value in row):
            continue
        yield dict(zip(headers, row))


def create_database():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    if DB_PATH.exists():
        DB_PATH.unlink()

    players_data = load_json(PLAYERS_JSON)
    ballon_data = load_json(BALLON_JSON)
    champions_wb = load_workbook(CHAMPIONS_XLSX, data_only=True)

    conn = sqlite3.connect(DB_PATH)
    conn.executescript(SCHEMA)

    conn.executemany(
        """
        INSERT INTO players (
          slug, name, matches, goals, assists,
          hub_matches, hub_goals, hub_assists,
          hub_matches_diff, hub_goals_diff, hub_assists_diff,
          collective_titles, individual_awards
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row["slug"],
                row["nome"],
                row["jogos"],
                row["gols"],
                row["assistencias"],
                row.get("jogos_hub"),
                row.get("gols_hub"),
                row.get("assistencias_hub"),
                row.get("dif_jogos_hub"),
                row.get("dif_gols_hub"),
                row.get("dif_assistencias_hub"),
                row.get("titulos_coletivos"),
                row.get("premios_individuais"),
            )
            for row in players_data["jogadores"]
        ],
    )

    conn.executemany(
        """
        INSERT INTO player_spells (
          id, player_slug, player_name, spell_order, team_name, team_full_name,
          team_type, matches, goals, assists, page_path, description
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row["id"],
                row["jogador_slug"],
                row["jogador"],
                row["ordem"],
                row["clube"],
                row["clube_nome_completo"],
                row["tipo"],
                row["jogos"],
                row["gols"],
                row["assistencias"],
                row["pagina"],
                row["descricao"],
            )
            for row in players_data["passagens"]
        ],
    )

    conn.executemany(
        """
        INSERT INTO player_competition_stats (
          spell_id, player_name, team_name, competition_name, matches, goals, assists
        )
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row["passagem_id"],
                row["jogador"],
                row["clube"],
                row["competicao"],
                row["jogos"],
                row["gols"],
                row["assistencias"],
            )
            for row in players_data["desempenho_por_competicao"]
        ],
    )

    conn.executemany(
        """
        INSERT INTO player_spell_honors (
          spell_id, player_name, team_name, category, honor_name, quantity
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row["passagem_id"],
                row["jogador"],
                row["clube"],
                row["categoria"],
                row["nome"],
                row["quantidade"],
            )
            for row in players_data["conquistas_por_passagem"]
        ],
    )

    conn.executemany(
        """
        INSERT INTO player_aggregate_honors (
          player_slug, player_name, category, honor_name, quantity
        )
        VALUES (?, ?, ?, ?, ?)
        """,
        [
            (
                row["jogador_slug"],
                row["jogador"],
                row["categoria"],
                row["nome"],
                row["quantidade"],
            )
            for row in players_data["conquistas_por_jogador"]
        ],
    )

    conn.executemany(
        """
        INSERT INTO championship_results (
          category, type, competition_name, season, champion_name, source
        )
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row["Categoria"],
                row["Tipo"],
                row["Competição"],
                str(row["Ano"]),
                row["Campeão"],
                row["Origem"],
            )
            for row in rows_from_sheet(champions_wb, "Base Consolidada")
        ],
    )

    conn.executemany(
        """
        INSERT INTO competition_champion_rankings (
          competition_name, champion_name, titles, first_year, last_year
        )
        VALUES (?, ?, ?, ?, ?)
        """,
        [
            (
                row["Competição"],
                row["Campeão"],
                row["Títulos"],
                row["Primeiro ano"],
                row["Último ano"],
            )
            for row in rows_from_sheet(champions_wb, "Ranking por Competição")
        ],
    )

    conn.executemany(
        """
        INSERT INTO club_relevant_titles (team_name, title_name, total, source)
        VALUES (?, ?, ?, ?)
        """,
        [
            (
                row["Clube/Seleção"],
                row["Título"],
                row["Total"],
                row["Fonte"],
            )
            for row in rows_from_sheet(champions_wb, "Títulos por Clube")
        ],
    )

    conn.executemany(
        """
        INSERT INTO ballon_rankings (year, position, player_name, podium, medal)
        VALUES (?, ?, ?, ?, ?)
        """,
        [
            (
                row["Ano"],
                row["Posição"],
                row["Jogador"],
                row["Pódio"],
                row["Medalha"],
            )
            for row in ballon_data["rankings_anuais"]
        ],
    )

    conn.executemany(
        """
        INSERT INTO ballon_player_stats (
          player_name, nominations, podiums, wins, seconds, thirds,
          best_position, best_position_years, first_nomination, last_nomination, history
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        [
            (
                row["Jogador"],
                row["Indicações"],
                row["Pódios"],
                row["Vitórias"],
                row["Vices"],
                row["Terceiros"],
                row["Melhor posição"],
                row["Ano(s) da melhor posição"],
                row["Primeira indicação"],
                row["Última indicação"],
                row["Histórico"],
            )
            for row in ballon_data["estatisticas_por_jogador"]
        ],
    )

    conn.commit()
    conn.close()
    print(DB_PATH)


if __name__ == "__main__":
    create_database()
