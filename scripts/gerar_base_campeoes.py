import json
from collections import Counter, defaultdict
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
INPUT_XLSX = Path(r"C:\Users\matheus\Downloads\campeoes_simplificado_1.xlsx")
SITE_DATA = ROOT / "data" / "site-dados-extraidos.json"
OUTPUT_XLSX = ROOT / "data" / "campeoes_completo_ate_2050.xlsx"


COMPETITION_MAP = {
    "Brasileirao": ("Ligas Nacionais", "Brasileirão", "Liga"),
    "Bundesliga": ("Ligas Nacionais", "Bundesliga (Alemanha)", "Liga"),
    "Premier League": ("Ligas Nacionais", "Premier League (Inglaterra)", "Liga"),
    "Serie A": ("Ligas Nacionais", "Serie A (Itália)", "Liga"),
    "Ligue 1": ("Ligas Nacionais", "Ligue 1 (França)", "Liga"),
    "LaLiga": ("Ligas Nacionais", "La Liga (Espanha)", "Liga"),
    "Copa do Brasil": ("Copas Nacionais", "Copa do Brasil", "Copa nacional"),
    "Supercopa do Brasil": ("Copas Nacionais", "Supercopa do Brasil", "Supercopa nacional"),
    "DFB Pokal": ("Copas Nacionais", "Copa da Alemanha (DFB-Pokal)", "Copa nacional"),
    "DFL Supercup": ("Copas Nacionais", "Supercopa da Alemanha (DFL Supercup)", "Supercopa nacional"),
    "Carabao Cup": ("Copas Nacionais", "Copa da Liga Inglesa (Carabao Cup)", "Copa nacional"),
    "FA Cup": ("Copas Nacionais", "Copa da Inglaterra (FA Cup)", "Copa nacional"),
    "Community Shield": ("Copas Nacionais", "Community Shield", "Supercopa nacional"),
    "Coppa Italia": ("Copas Nacionais", "Copa da Itália (Coppa Italia)", "Copa nacional"),
    "Supercoppa Italiana": ("Copas Nacionais", "Supercopa da Itália", "Supercopa nacional"),
    "Coupe de France": ("Copas Nacionais", "Copa da França", "Copa nacional"),
    "Supercopa da Franca": ("Copas Nacionais", "Supercopa da França", "Supercopa nacional"),
    "Copa del Rey": ("Copas Nacionais", "Copa del Rey (Espanha)", "Copa nacional"),
    "Supercopa da Espanha": ("Copas Nacionais", "Supercopa da Espanha", "Supercopa nacional"),
    "Champions League": ("Competições Europeias", "Champions League", "Continental clubes"),
    "Europa League": ("Competições Europeias", "Europa League", "Continental clubes"),
    "Conference League": ("Competições Europeias", "Conference League", "Continental clubes"),
    "Supercopa da Europa": ("Competições Europeias", "Supercopa da Europa", "Supercopa continental"),
    "Libertadores": ("América do Sul", "Copa Libertadores", "Continental clubes"),
    "Sul-Americana": ("América do Sul", "Copa Sul-Americana", "Continental clubes"),
    "Recopa Sul-Americana": ("América do Sul", "Recopa Sudamericana", "Supercopa continental"),
    "Copa do Mundo": ("Copa do Mundo e Continentais", "Copa do Mundo FIFA", "Seleções"),
    "Eurocopa": ("Copa do Mundo e Continentais", "Eurocopa", "Seleções"),
    "Copa America": ("Copa do Mundo e Continentais", "Copa América", "Seleções"),
    "Mundial de Clubes": ("Copa do Mundo e Continentais", "Mundial de Clubes FIFA", "Mundial clubes"),
    "Super Mundial": ("Copa do Mundo e Continentais", "Supermundial", "Mundial clubes"),
}


NAME_MAP = {
    "Sao Paulo": "São Paulo",
    "Gremio": "Grêmio",
    "Ceara": "Ceará",
    "Atletico Mineiro": "Atlético Mineiro",
    "Atletico de Madrid": "Atlético de Madrid",
    "Barcelona de Guayaquil": "Barcelona de Guayaquil",
    "Frankfurt": "Eintracht Frankfurt",
    "Bayern": "Bayern de Munique",
    "Leipzig": "RB Leipzig",
    "PSG": "Paris Saint-Germain",
    "Spain": "Espanha",
    "Italy": "Itália",
    "Netherlands": "Holanda",
    "France": "França",
    "Brazil": "Brasil",
    "Germany": "Alemanha",
    "Uruguay": "Uruguai",
    "England": "Inglaterra",
    "Bolivia": "Bolívia",
    "Mexico": "México",
}


TITLE_MAP = {
    "Brasileirao": "Brasileirão",
    "Supercopa da Franca": "Supercopa da França",
    "Copa America": "Copa América",
    "Copa do Mundo": "Copa do Mundo FIFA",
    "DFB Pokal": "Copa da Alemanha (DFB-Pokal)",
    "DFL Supercup": "Supercopa da Alemanha (DFL Supercup)",
    "Coppa Italia": "Copa da Itália (Coppa Italia)",
    "Supercoppa Italiana": "Supercopa da Itália",
}


RANKING_OVERRIDES = {
    ("Champions League", "Ajax"): 5,
}


def canon_name(name):
    if name is None:
        return ""
    return NAME_MAP.get(str(name), str(name))


def canon_title(title):
    if title is None:
        return ""
    return TITLE_MAP.get(str(title), str(title))


def copy_row_style(ws, source_row, target_row):
    for col in range(1, ws.max_column + 1):
        src = ws.cell(source_row, col)
        dst = ws.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)


def clear_generated_sheets(wb):
    for sheet_name in [
        "Base Consolidada",
        "Ranking por Competição",
        "Títulos por Clube",
        "Resumo",
    ]:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 3, 12), 42)


def existing_rows_by_sheet(wb):
    rows = []
    for sheet_name in [
        "Competições Europeias",
        "América do Sul",
        "Copa do Mundo e Continentais",
        "Ligas Nacionais",
        "Copas Nacionais",
    ]:
        ws = wb[sheet_name]
        for comp, year, champion in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            if not comp or not year or not champion:
                continue
            rows.append(
                {
                    "Categoria": sheet_name,
                    "Tipo": "Pré-simulação",
                    "Competição": str(comp),
                    "Ano": str(year),
                    "Campeão": canon_name(champion),
                    "Origem": "Tabela inicial",
                }
            )
    return rows


def append_simulation_rows(wb, competitions):
    generated = []
    for competition in competitions:
        source_name = competition["name"]
        if source_name not in COMPETITION_MAP:
            continue

        sheet_name, display_name, comp_type = COMPETITION_MAP[source_name]
        ws = wb[sheet_name]
        existing_keys = {
            (str(row[0]), str(row[1]), canon_name(row[2]))
            for row in ws.iter_rows(min_row=2, max_col=3, values_only=True)
            if row[0] and row[1] and row[2]
        }

        for year, champion in sorted(competition["winners"].items(), key=lambda item: int(item[0])):
            year = int(year)
            if year > 2050:
                continue

            display_champion = canon_name(champion)
            key = (display_name, str(year), display_champion)
            if key not in existing_keys:
                target_row = ws.max_row + 1
                copy_row_style(ws, max(2, ws.max_row), target_row)
                ws.cell(target_row, 1).value = display_name
                ws.cell(target_row, 2).value = str(year)
                ws.cell(target_row, 3).value = display_champion
                existing_keys.add(key)

            generated.append(
                {
                    "Categoria": sheet_name,
                    "Tipo": comp_type,
                    "Competição": display_name,
                    "Ano": str(year),
                    "Campeão": display_champion,
                    "Origem": "Site / simulação",
                }
            )
    return generated


def add_summary_sheets(wb, all_rows, team_data):
    base = wb.create_sheet("Base Consolidada")
    base_rows = [["Categoria", "Tipo", "Competição", "Ano", "Campeão", "Origem"]]
    base_rows.extend(
        [
            [
                row["Categoria"],
                row["Tipo"],
                row["Competição"],
                row["Ano"],
                row["Campeão"],
                row["Origem"],
            ]
            for row in sorted(all_rows, key=lambda item: (item["Categoria"], item["Competição"], item["Ano"]))
        ]
    )
    for row in base_rows:
        base.append(row)
    style_sheet(base)

    ranking = wb.create_sheet("Ranking por Competição")
    ranking.append(["Competição", "Campeão", "Títulos", "Primeiro ano", "Último ano"])
    grouped = defaultdict(list)
    for row in all_rows:
        grouped[row["Competição"]].append(row)

    for competition_name in sorted(grouped):
        counts = Counter(row["Campeão"] for row in grouped[competition_name])
        years_by_team = defaultdict(list)
        for row in grouped[competition_name]:
            year_value = str(row["Ano"]).split("–")[0].split("-")[0]
            try:
                years_by_team[row["Campeão"]].append(int(year_value))
            except ValueError:
                pass

        for (override_competition, override_team), override_total in RANKING_OVERRIDES.items():
            if competition_name == override_competition:
                counts[override_team] = max(counts.get(override_team, 0), override_total)

        for team, total in sorted(counts.items(), key=lambda item: (-item[1], item[0])):
            years = years_by_team[team]
            ranking.append([competition_name, team, total, min(years) if years else "", max(years) if years else ""])
    style_sheet(ranking)

    titles = wb.create_sheet("Títulos por Clube")
    titles.append(["Clube/Seleção", "Título", "Total", "Fonte"])
    title_rows = []
    for team, data in team_data.items():
        for title, total in (data.get("titles") or {}).items():
            title_rows.append([canon_name(team), canon_title(title), total, "teamData do site"])

    for row in sorted(title_rows, key=lambda item: (item[0], item[1])):
        titles.append(row)
    style_sheet(titles)

    resumo = wb.create_sheet("Resumo", 0)
    sim_rows = [row for row in all_rows if row["Origem"] == "Site / simulação"]
    resumo_rows = [
        ["Indicador", "Valor"],
        ["Competições com dados da simulação", len({row["Competição"] for row in sim_rows})],
        ["Linhas adicionadas da simulação", len(sim_rows)],
        ["Período da simulação usado", "2026-2050"],
        ["Clubes/seleções com títulos relevantes", len(team_data)],
        ["Fonte principal", "assets/js/dados.js"],
    ]
    for row in resumo_rows:
        resumo.append(row)
    style_sheet(resumo)


def main():
    with SITE_DATA.open("r", encoding="utf-8") as file:
        site_data = json.load(file)

    wb = load_workbook(INPUT_XLSX)
    clear_generated_sheets(wb)
    generated_rows = append_simulation_rows(wb, site_data["competitions"])
    all_rows = existing_rows_by_sheet(wb)

    # Replace duplicated generated rows in the consolidated base with the normalized rows built above.
    all_rows = [row for row in all_rows if not (row["Ano"].isdigit() and int(row["Ano"]) >= 2026)]
    all_rows.extend(generated_rows)

    add_summary_sheets(wb, all_rows, site_data["teamData"])

    for sheet_name in wb.sheetnames:
        style_sheet(wb[sheet_name])

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
