import html
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
PLAYER_DIR = ROOT / "pages" / "jogadores"
CLUB_DIR = ROOT / "pages" / "clubes"
OUTPUT_XLSX = ROOT / "data" / "jogadores_carreiras.xlsx"
OUTPUT_JSON = ROOT / "data" / "jogadores_carreiras.json"

PLAYER_FILES = [
    "garcia.html",
    "fonseca.html",
    "escobar.html",
    "salazar.html",
    "guarita-fonseca.html",
    "moreno-escobar.html",
]

CLUB_NAME_MAP = {
    "Ajax": "Ajax",
    "Barça": "Barcelona",
    "Bayern": "Bayern de Munique",
    "Bolívia": "Bolívia",
    "Bolivia": "Bolívia",
    "City": "Manchester City",
    "Corinthians": "Corinthians",
    "Grêmio": "Grêmio",
    "Marseille": "Olympique de Marseille",
    "Milan": "Milan",
    "Nacional de Potosí": "Nacional de Potosí",
    "Oriente": "Oriente Petrolero",
    "Oriente Petrolero": "Oriente Petrolero",
    "PSG": "Paris Saint-Germain",
    "Paris Saint-Germain": "Paris Saint-Germain",
    "Seleção da Bolívia": "Bolívia",
    "The Strongest": "The Strongest",
    "Vasco": "Vasco",
}

COMPETITION_NAME_MAP = {
    "Brasileirão B": "Brasileirão Série B",
    "Carabao": "Carabao Cup",
    "Champions": "Champions League",
    "Copa": "Copa do Mundo",
    "Copa da Itália": "Copa da Itália (Coppa Italia)",
    "Copa do Rey": "Copa del Rey",
    "Coupe de France": "Copa da França",
    "Copa da Franca": "Copa da França",
    "DFB Pokal": "Copa da Alemanha (DFB-Pokal)",
    "DFB-Pokal": "Copa da Alemanha (DFB-Pokal)",
    "Emirates": "FA Cup",
    "Intercontinental": "Mundial de Clubes",
    "Liga A": "Serie A",
    "Mundial": "Mundial de Clubes",
    "Recopa": "Recopa Sul-Americana",
    "Sudamericana": "Sul-Americana",
    "SuperCup": "Supercopa da Europa",
    "SuperMundial": "Super Mundial",
    "Supercup": "Supercopa da Europa",
    "Supermundial": "Super Mundial",
    "Supercopa": "Supercopa Rei",
    "Supercopa da Alemanha": "Supercopa da Alemanha (DFL Supercup)",
    "Supercopa da França": "Supercopa da França",
    "Supercopa da Franca": "Supercopa da França",
    "Supercopa da Inglaterra": "Community Shield",
}

TROPHY_NAME_MAP = {
    "Brasileirão B": "Brasileirão Série B",
    "Copa": "Copa do Mundo",
    "Copa da Itália": "Copa da Itália (Coppa Italia)",
    "DFB Pokal": "Copa da Alemanha (DFB-Pokal)",
    "DFB-Pokal": "Copa da Alemanha (DFB-Pokal)",
    "Intercontinental": "Mundial de Clubes",
    "Recopa": "Recopa Sul-Americana",
    "Supercup": "Supercopa da Europa",
    "SuperCup": "Supercopa da Europa",
    "Super Mundial": "Super Mundial",
    "Supercopa Rei": "Supercopa Rei",
    "Supercopa da Alemanha": "Supercopa da Alemanha (DFL Supercup)",
    "Copa da Franca": "Copa da França",
    "Supercopa da Franca": "Supercopa da França",
}


def clean_text(value):
    value = re.sub(r"<[^>]+>", "", value or "")
    value = html.unescape(value)
    value = re.sub(r"\s+", " ", value).strip()
    if any(token in value for token in ["Ã", "Â", "â"]):
        try:
            value = value.encode("latin1").decode("utf-8")
        except UnicodeError:
            pass
    return value


def int_from_count(value):
    match = re.search(r"\d+", str(value or ""))
    return int(match.group(0)) if match else 0


def slugify(value):
    value = value.lower()
    replacements = {
        "á": "a",
        "à": "a",
        "ã": "a",
        "â": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    for old, new in replacements.items():
        value = value.replace(old, new)
    return re.sub(r"[^a-z0-9]+", "-", value).strip("-")


def normalize_for_match(value):
    replacements = {
        "á": "a",
        "à": "a",
        "ã": "a",
        "â": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    value = clean_text(value).lower()
    for old, new in replacements.items():
        value = value.replace(old, new)
    return value


def canonical_club_name(value):
    value = clean_text(value)
    return CLUB_NAME_MAP.get(value, value)


def canonical_competition_name(value):
    value = clean_text(value)
    return COMPETITION_NAME_MAP.get(value, value)


def canonical_trophy_name(value):
    value = clean_text(value)
    return TROPHY_NAME_MAP.get(value, value)


def extract_cards(content, card_class):
    pattern = rf'<article class="{card_class}"[^>]*>(.*?)</article>'
    return re.findall(pattern, content, re.S)


def extract_metric_cards(content, card_class):
    rows = {}
    for card in extract_cards(content, card_class):
        label = clean_text(re.search(r'__label">(.+?)</div>', card, re.S).group(1))
        value = clean_text(re.search(r'__value">(.+?)</div>', card, re.S).group(1))
        rows[label] = int_from_count(value)
    return rows


def metric_value(rows, *labels):
    for label in labels:
        if label in rows:
            return rows[label]
    normalized_labels = [normalize_for_match(label) for label in labels]
    for key, value in rows.items():
        normalized_key = normalize_for_match(key)
        if normalized_key in normalized_labels:
            return value
    if any("assist" in label for label in normalized_labels):
        for key, value in rows.items():
            if "assist" in normalize_for_match(key):
                return value
    if any("titulo" in label for label in normalized_labels):
        for key, value in rows.items():
            normalized_key = normalize_for_match(key)
            if normalized_key.startswith("t") and "coletivos" in normalized_key:
                return value
    if any("premio" in label for label in normalized_labels):
        for key, value in rows.items():
            normalized_key = normalize_for_match(key)
            if normalized_key.startswith("pr") and "individuais" in normalized_key:
                return value
    return 0


def extract_awards_by_section(content):
    sections = re.findall(r'<section class="[^"]*\bsection\b[^"]*"[^>]*>(.*?)</section>', content, re.S)
    awards = []
    for section in sections:
        title_match = re.search(r'<h2 class="section__title">(.+?)</h2>', section, re.S)
        if not title_match:
            continue
        title = clean_text(title_match.group(1)).lower()
        normalized_title = normalize_for_match(title)
        if "titulo" in normalized_title or "conquista" in normalized_title:
            category = "Título coletivo"
        elif "premio" in normalized_title or "reconhecimento" in normalized_title:
            category = "Prêmio individual"
        else:
            continue

        for card in extract_cards(section, "award-card"):
            name_match = re.search(r'award-card__name">(.+?)</div>', card, re.S)
            count_match = re.search(r'award-card__count">(.+?)</div>', card, re.S)
            if not name_match or not count_match:
                continue
            awards.append(
                {
                    "categoria": category,
                    "nome": canonical_trophy_name(name_match.group(1)),
                    "quantidade": int_from_count(clean_text(count_match.group(1))),
                }
            )
    return awards


def parse_player_page(path):
    content = path.read_text(encoding="utf-8")
    name = clean_text(re.search(r'<h1 class="hero__title">(.+?)</h1>', content, re.S).group(1))
    metrics = extract_metric_cards(content, "metric-card")
    summary = extract_metric_cards(content, "summary-card")

    clubs = []
    for match in re.finditer(r'<a class="club-card[^"]*" href="(?P<href>[^"]+)"[^>]*>(?P<body>.*?)</a>', content, re.S):
        body = match.group("body")
        name_match = re.search(r'club-card__name">(.+?)</h3>', body, re.S)
        caption_match = re.search(r'club-card__caption">(.+?)</p>', body, re.S)
        clubs.append(
            {
                "href": match.group("href").replace("../clubes/", ""),
                "nome": clean_text(name_match.group(1)) if name_match else "",
                "descricao": clean_text(caption_match.group(1)) if caption_match else "",
            }
        )

    return {
        "slug": path.stem,
        "nome": name,
        "jogos": metrics.get("Jogos", 0),
        "gols": metrics.get("Gols", 0),
        "assistencias": metric_value(metrics, "Assistências", "AssistÃªncias", "Assistencias"),
        "titulos_coletivos": summary.get("Títulos coletivos", summary.get("TÃ­tulos coletivos", 0)),
        "premios_individuais": summary.get("Prêmios individuais", summary.get("PrÃªmios individuais", 0)),
        "clubes": clubs,
        "conquistas": extract_awards_by_section(content),
    }


def parse_club_page(path):
    content = path.read_text(encoding="utf-8")
    title_match = re.search(r"<title>(.*?)</title>", content, re.S)
    player_name = clean_text(title_match.group(1).split("|")[0]) if title_match else ""

    club_title = re.search(r'<h1 class="club-hero__title">(.+?)</h1>', content, re.S)
    club_name = clean_text(club_title.group(1)) if club_title else path.stem
    alt_match = re.search(r'<div class="club-hero__logo"><img [^>]*alt="([^"]+)"', content, re.S)
    full_club_name = clean_text(alt_match.group(1)) if alt_match else club_name
    club_name = canonical_club_name(club_name)
    full_club_name = canonical_club_name(full_club_name)

    summary = extract_metric_cards(content, "summary-card")

    competition_stats = []
    table_match = re.search(r"<tbody>(.*?)</tbody>", content, re.S)
    if table_match:
        for row in re.findall(r"<tr[^>]*>(.*?)</tr>", table_match.group(1), re.S):
            cells = [clean_text(cell) for cell in re.findall(r"<td>(.*?)</td>", row, re.S)]
            if len(cells) != 4 or cells[0].lower() == "total":
                continue
            competition_stats.append(
                {
                    "competicao": canonical_competition_name(cells[0]),
                    "jogos": int_from_count(cells[1]),
                    "gols": int_from_count(cells[2]),
                    "assistencias": int_from_count(cells[3]),
                }
            )

    trophies = []
    for card in extract_cards(content, "trophy-card"):
        name_match = re.search(r'trophy-card__name">(.+?)</div>', card, re.S)
        count_match = re.search(r'trophy-card__count">(.+?)</div>', card, re.S)
        if not name_match or not count_match:
            continue
        name = canonical_trophy_name(name_match.group(1))
        category = "Prêmio individual" if re.search(r"Melhor|Artilheiro|Bola de Ouro|Chuteira", name, re.I) else "Título coletivo"
        trophies.append(
            {
                "categoria": category,
                "nome": name,
                "quantidade": int_from_count(clean_text(count_match.group(1))),
            }
        )

    return {
        "arquivo": path.name,
        "jogador": player_name,
        "jogador_slug": slugify(player_name),
        "clube": club_name,
        "clube_nome_completo": full_club_name,
        "jogos": summary.get("Jogos", 0),
        "gols": summary.get("Gols", 0),
        "assistencias": metric_value(summary, "Assistências", "AssistÃªncias", "Assistencias"),
        "estatisticas_competicao": competition_stats,
        "conquistas": trophies,
    }


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_len + 3, 12), 62)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_payload():
    players = [parse_player_page(PLAYER_DIR / file_name) for file_name in PLAYER_FILES]
    club_pages = [parse_club_page(path) for path in sorted(CLUB_DIR.glob("*.html"))]

    player_by_name = {player["nome"]: player for player in players}
    player_name_to_slug = {player["nome"]: player["slug"] for player in players}
    for page in club_pages:
        page["jogador_slug"] = player_name_to_slug.get(page["jogador"], page["jogador_slug"])

    spells = []
    spell_lookup = {}
    for player in players:
        for order, club in enumerate(player["clubes"], start=1):
            page = next((item for item in club_pages if item["arquivo"] == club["href"]), None)
            if not page:
                continue
            spell_id = f'{player["slug"]}-{slugify(page["clube_nome_completo"])}-{order}'
            row = {
                "id": spell_id,
                "jogador": player["nome"],
                "jogador_slug": player["slug"],
                "ordem": order,
                "clube": page["clube"],
                "clube_nome_completo": page["clube_nome_completo"],
                "tipo": "Seleção" if "Bolívia" in page["clube_nome_completo"] or "Bolivia" in page["clube_nome_completo"] else "Clube",
                "jogos": page["jogos"],
                "gols": page["gols"],
                "assistencias": page["assistencias"],
                "pagina": f'pages/clubes/{page["arquivo"]}',
                "descricao": club["descricao"],
            }
            spells.append(row)
            spell_lookup[page["arquivo"]] = spell_id

    competition_rows = []
    spell_titles = []
    for page in club_pages:
        spell_id = spell_lookup.get(page["arquivo"])
        if not spell_id:
            continue
        for item in page["estatisticas_competicao"]:
            competition_rows.append({"passagem_id": spell_id, "jogador": page["jogador"], "clube": page["clube_nome_completo"], **item})
        for trophy in page["conquistas"]:
            spell_titles.append({"passagem_id": spell_id, "jogador": page["jogador"], "clube": page["clube_nome_completo"], **trophy})

    player_awards = []
    for player in players:
        for item in player["conquistas"]:
            player_awards.append({"jogador": player["nome"], "jogador_slug": player["slug"], **item})

    checks = []
    spells_by_player = defaultdict(list)
    for spell in spells:
        spells_by_player[spell["jogador"]].append(spell)
    for player in players:
        player_spells = spells_by_player[player["nome"]]
        real_jogos = sum(item["jogos"] for item in player_spells)
        real_gols = sum(item["gols"] for item in player_spells)
        real_assistencias = sum(item["assistencias"] for item in player_spells)
        player["jogos_hub"] = player["jogos"]
        player["gols_hub"] = player["gols"]
        player["assistencias_hub"] = player["assistencias"]
        player["dif_jogos_hub"] = real_jogos - player["jogos_hub"]
        player["dif_gols_hub"] = real_gols - player["gols_hub"]
        player["dif_assistencias_hub"] = real_assistencias - player["assistencias_hub"]
        player["jogos"] = real_jogos
        player["gols"] = real_gols
        player["assistencias"] = real_assistencias

        checks.append(
            {
                "jogador": player["nome"],
                "jogos_hub": player["jogos_hub"],
                "jogos_reais": real_jogos,
                "dif_jogos": player["dif_jogos_hub"],
                "gols_hub": player["gols_hub"],
                "gols_reais": real_gols,
                "dif_gols": player["dif_gols_hub"],
                "assistencias_hub": player["assistencias_hub"],
                "assistencias_reais": real_assistencias,
                "dif_assistencias": player["dif_assistencias_hub"],
                "status": "OK"
                if (
                    player["dif_jogos_hub"] == 0
                    and player["dif_gols_hub"] == 0
                    and player["dif_assistencias_hub"] == 0
                )
                else "Total real aplicado",
            }
        )

    return {
        "jogadores": players,
        "passagens": spells,
        "desempenho_por_competicao": competition_rows,
        "conquistas_por_passagem": spell_titles,
        "conquistas_por_jogador": player_awards,
        "checks": checks,
    }


def write_workbook(payload):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Indicador", "Valor"])
    ws.append(["Jogadores", len(payload["jogadores"])])
    ws.append(["Passagens por clube/seleção", len(payload["passagens"])])
    ws.append(["Linhas de desempenho por competição", len(payload["desempenho_por_competicao"])])
    ws.append(["Conquistas por passagem", len(payload["conquistas_por_passagem"])])
    ws.append(["Conquistas agregadas por jogador", len(payload["conquistas_por_jogador"])])

    sheets = [
        (
            "Jogadores",
            [
                "slug",
                "nome",
                "jogos",
                "gols",
                "assistencias",
                "jogos_hub",
                "gols_hub",
                "assistencias_hub",
                "dif_jogos_hub",
                "dif_gols_hub",
                "dif_assistencias_hub",
                "titulos_coletivos",
                "premios_individuais",
            ],
            payload["jogadores"],
        ),
        (
            "Passagens",
            ["id", "jogador_slug", "jogador", "ordem", "clube", "clube_nome_completo", "tipo", "jogos", "gols", "assistencias", "pagina", "descricao"],
            payload["passagens"],
        ),
        (
            "Desempenho Competição",
            ["passagem_id", "jogador", "clube", "competicao", "jogos", "gols", "assistencias"],
            payload["desempenho_por_competicao"],
        ),
        (
            "Conquistas Passagem",
            ["passagem_id", "jogador", "clube", "categoria", "nome", "quantidade"],
            payload["conquistas_por_passagem"],
        ),
        (
            "Conquistas Jogador",
            ["jogador_slug", "jogador", "categoria", "nome", "quantidade"],
            payload["conquistas_por_jogador"],
        ),
        (
            "Checks",
            [
                "jogador",
                "jogos_hub",
                "jogos_reais",
                "dif_jogos",
                "gols_hub",
                "gols_reais",
                "dif_gols",
                "assistencias_hub",
                "assistencias_reais",
                "dif_assistencias",
                "status",
            ],
            payload["checks"],
        ),
    ]

    for sheet_name, headers, rows in sheets:
        sheet = wb.create_sheet(sheet_name)
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])

    for sheet in wb.worksheets:
        style_sheet(sheet)

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)


def main():
    payload = build_payload()
    write_workbook(payload)
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(OUTPUT_XLSX)
    print(OUTPUT_JSON)


if __name__ == "__main__":
    main()
